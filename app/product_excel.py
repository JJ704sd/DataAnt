from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.product_models import ProductRecord, ProductStatus


PRODUCT_COLUMNS: list[str] = [
    "product_id",
    "source_site",
    "product_url",
    "name",
    "category",
    "description",
    "primary_image_url",
    "current_price",
    "original_price",
    "currency",
    "brand",
    "variant_count",
    "status",
    "error_message",
    "collected_at",
]


@dataclass(frozen=True, slots=True)
class ProductWriteReceipt:
    product_ids: tuple[str, ...]
    row_count: int
    bytes_written: int


class ProductExcel:
    """Workbook I/O for the bounded product gallery.

    ``write`` produces a brand-new ``products`` sheet using the fixed 15-column
    contract. ``read`` reverses the workbook into ``ProductRecord`` instances
    and ``merge_existing`` upserts by ``product_id`` so re-runs do not produce
    duplicates. The atomic temp-file + rename swap is the caller's
    responsibility; this class only writes the caller-supplied path.
    """

    @staticmethod
    def _workbook_for(path: Path) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "products"
        sheet.append(list(PRODUCT_COLUMNS))
        return workbook, sheet

    @classmethod
    def write(
        cls,
        path: Path,
        records: list[ProductRecord],
        *,
        primitive_rows: Sequence[Mapping[str, object]] | None = None,
    ) -> ProductWriteReceipt:
        path.parent.mkdir(parents=True, exist_ok=True)
        rows: tuple[Mapping[str, object], ...] = (
            tuple(record.to_primitive() for record in records)
            if primitive_rows is None
            else tuple(primitive_rows)
        )
        if len(rows) != len(records):
            raise ValueError("primitive rows must match records")

        workbook, sheet = cls._workbook_for(path)
        for row in rows:
            sheet.append([row[column] for column in PRODUCT_COLUMNS])
        workbook.save(path)
        return ProductWriteReceipt(
            product_ids=tuple(record.product_id for record in records),
            row_count=len(records),
            bytes_written=path.stat().st_size,
        )

    @classmethod
    def read(cls, path: Path) -> list[ProductRecord]:
        workbook = load_workbook(path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        if header != PRODUCT_COLUMNS:
            raise ValueError(
                "product workbook schema does not match the 15-column contract"
            )
        records: list[ProductRecord] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            records.append(cls._record_from_row(row))
        return records

    @classmethod
    def merge_existing(
        cls, path: Path, new_records: list[ProductRecord]
    ) -> list[ProductRecord]:
        merged: list[ProductRecord] = []
        seen: set[str] = set()
        for record in cls.read(path):
            if record.product_id in seen:
                continue
            seen.add(record.product_id)
            merged.append(record)
        for record in new_records:
            if record.product_id in seen:
                for index, existing_record in enumerate(merged):
                    if existing_record.product_id == record.product_id:
                        merged[index] = record
                        break
            else:
                merged.append(record)
                seen.add(record.product_id)
        return merged

    @staticmethod
    def _row_for(record: ProductRecord) -> list[object]:
        payload = record.to_primitive()
        return [payload[column] for column in PRODUCT_COLUMNS]

    @staticmethod
    def _record_from_row(row: tuple[object, ...]) -> ProductRecord:
        values = dict(zip(PRODUCT_COLUMNS, row))
        current_price = values.get("current_price")
        original_price = values.get("original_price")
        status_value = values.get("status")
        return ProductRecord(
            product_id=str(values.get("product_id", "")),
            source_site=str(values.get("source_site", "")),
            product_url=str(values.get("product_url", "")),
            name=str(values.get("name", "") or ""),
            category=str(values.get("category", "") or ""),
            description=str(values.get("description", "") or ""),
            primary_image_url=str(values.get("primary_image_url", "") or ""),
            current_price=(
                Decimal(str(current_price)) if current_price not in (None, "") else None
            ),
            original_price=(
                Decimal(str(original_price))
                if original_price not in (None, "")
                else None
            ),
            currency=str(values.get("currency", "") or ""),
            brand=str(values.get("brand", "") or ""),
            variant_count=int(values.get("variant_count") or 0),
            status=ProductStatus(str(status_value)) if status_value else ProductStatus.UNEXPECTED_ERROR,
            error_message=str(values.get("error_message", "") or ""),
            collected_at=str(values.get("collected_at", "") or ""),
        )
