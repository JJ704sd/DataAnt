from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import app.excel_store as excel_store_module
from app.excel_store import COLUMNS, ExcelStore, OutputLockedError
from app.models import MatchMethod, MovieResult, Status


def success(task_id: str, title: str, rating: float | None = 9.0) -> MovieResult:
    return MovieResult(
        task_id=task_id,
        query=title,
        query_year="1994",
        matched_title=title,
        matched_year="1994",
        director="Director One / Director Two",
        rating=rating,
        detail_url="https://movie.douban.com/subject/1/",
        match_method=MatchMethod.RULE_EXACT,
        status=Status.SUCCESS,
        collected_at="2026-07-15T12:00:00+08:00",
    )


def test_upsert_creates_exact_schema_and_serializes_values(tmp_path: Path) -> None:
    path = tmp_path / "result.xlsx"
    ExcelStore(path).upsert(success("a", "Movie", rating=None))
    workbook = load_workbook(path)
    assert workbook.active.title == "movies"
    rows = list(workbook.active.values)
    assert list(rows[0]) == COLUMNS
    assert rows[1][0] == "a"
    assert rows[1][6] is None
    assert rows[1][8] == "RULE_EXACT"
    assert rows[1][9] == "SUCCESS"


def test_upsert_replaces_same_task_without_duplicate_row(tmp_path: Path) -> None:
    path = tmp_path / "result.xlsx"
    store = ExcelStore(path)
    store.upsert(success("a", "First"))
    store.upsert(success("a", "Updated"))
    rows = list(load_workbook(path).active.values)
    assert len(rows) == 2
    assert rows[1][1] == "Updated"


def test_status_index_survives_restart(tmp_path: Path) -> None:
    path = tmp_path / "result.xlsx"
    ExcelStore(path).upsert(success("a", "Movie"))
    assert ExcelStore(path).status_by_task_id() == {"a": Status.SUCCESS}


def test_missing_workbook_has_empty_status_index(tmp_path: Path) -> None:
    assert ExcelStore(tmp_path / "missing.xlsx").status_by_task_id() == {}


def test_existing_wrong_schema_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "wrong.xlsx"
    workbook = Workbook()
    workbook.active.append(["wrong"])
    workbook.save(path)
    with pytest.raises(ValueError, match="12-column contract"):
        ExcelStore(path).status_by_task_id()


def test_permission_error_preserves_target_and_temporary_file(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "result.xlsx"
    ExcelStore(path).upsert(success("a", "Original"))
    original_bytes = path.read_bytes()

    def locked_replace(source: Path, target: Path) -> None:
        raise PermissionError("locked")

    monkeypatch.setattr(excel_store_module.os, "replace", locked_replace)
    with pytest.raises(OutputLockedError, match="Close Excel and retry"):
        ExcelStore(path).upsert(success("a", "Updated"))

    assert path.read_bytes() == original_bytes
    assert path.with_suffix(".xlsx.tmp").is_file()
