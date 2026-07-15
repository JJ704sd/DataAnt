from __future__ import annotations

import os
from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType

from app.models import MovieResult, Status


COLUMNS = [
    "task_id",
    "query",
    "query_year",
    "matched_title",
    "matched_year",
    "director",
    "rating",
    "detail_url",
    "match_method",
    "status",
    "error_message",
    "collected_at",
]


class OutputLockedError(OSError):
    pass


class ExcelStore:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def _workbook(self) -> WorkbookType:
        if self.path.exists():
            workbook = load_workbook(self.path)
            header = [cell.value for cell in workbook.active[1]]
            if header != COLUMNS:
                raise ValueError(
                    "Existing workbook schema does not match the 12-column contract"
                )
            return workbook
        workbook = Workbook()
        workbook.active.title = "movies"
        workbook.active.append(COLUMNS)
        return workbook

    def status_by_task_id(self) -> dict[str, Status]:
        if not self.path.exists():
            return {}
        workbook = self._workbook()
        return {
            str(row[0]): Status(str(row[9]))
            for row in workbook.active.iter_rows(min_row=2, values_only=True)
        }

    def upsert(self, result: MovieResult) -> None:
        workbook = self._workbook()
        sheet = workbook.active
        row_number = next(
            (
                row[0].row
                for row in sheet.iter_rows(min_row=2)
                if row[0].value == result.task_id
            ),
            sheet.max_row + 1,
        )
        values: dict[str, Any] = asdict(result)
        for column_number, name in enumerate(COLUMNS, start=1):
            value = values[name]
            sheet.cell(
                row=row_number,
                column=column_number,
                value=getattr(value, "value", value),
            )
        temporary = self.path.with_suffix(self.path.suffix + ".tmp")
        try:
            workbook.save(temporary)
            os.replace(temporary, self.path)
        except PermissionError as exc:
            raise OutputLockedError(f"Close Excel and retry: {self.path}") from exc
