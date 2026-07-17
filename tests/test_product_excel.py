from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.product_excel import PRODUCT_COLUMNS, ProductExcel, ProductWriteReceipt
from app.product_models import ProductRecord


def test_render_creates_products_sheet_with_exact_columns(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    ProductExcel.write(path, [ProductRecord.success_fixture("1")])
    workbook = load_workbook(path)
    assert workbook.active.title == "products"
    rows = list(workbook.active.values)
    assert list(rows[0]) == PRODUCT_COLUMNS
    assert rows[1][0] == "1"
    assert rows[1][12] == "SUCCESS"


def test_write_accepts_shared_rows_and_returns_receipt(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    record = ProductRecord.success_fixture("1")

    receipt = ProductExcel.write(
        path,
        [record],
        primitive_rows=(record.to_primitive(),),
    )

    assert isinstance(receipt, ProductWriteReceipt)
    assert receipt.product_ids == ("1",)
    assert receipt.row_count == 1
    assert receipt.bytes_written == path.stat().st_size


def test_write_rejects_shared_row_length_mismatch(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    with pytest.raises(ValueError, match="primitive rows"):
        ProductExcel.write(
            path,
            [ProductRecord.success_fixture("1")],
            primitive_rows=(),
        )


def test_merge_replaces_product_id_without_duplicate(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    ProductExcel.write(path, [ProductRecord.success_fixture("1")])
    updated = ProductRecord.success_fixture("1")
    merged = ProductExcel.merge_existing(path, [updated])
    assert [record.product_id for record in merged] == ["1"]


def test_wrong_existing_schema_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    workbook = Workbook()
    workbook.active.append(["wrong"])
    workbook.save(path)
    with pytest.raises(ValueError, match="product workbook schema"):
        ProductExcel.read(path)
