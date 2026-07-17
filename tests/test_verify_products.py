"""Offline contract checks for the three-file product bundle.

The verifier must read ``products.xlsx``, ``products.json`` and
``gallery.html`` from a target directory and confirm they correspond
to the same bounded result set. All inputs come from the in-process
``ProductOutputBundle`` so the suite stays fully offline.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.product_models import (
    ProductCollection,
    ProductListing,
    ProductPage,
    ProductRecord,
    ProductStatus,
)
from app.product_output_bundle import ProductOutputBundle
from app.product_runner import ProductRunner
from scripts.verify_products import (
    ProductBundleContractError,
    verify_product_bundle,
)


BUNDLE_FILES: tuple[str, ...] = (
    "products.xlsx",
    "products.json",
    "gallery.html",
)


EXCEL_COLUMNS: tuple[str, ...] = (
    "product_id", "source_site", "product_url", "name", "category",
    "description", "primary_image_url", "current_price", "original_price",
    "currency", "brand", "variant_count", "status", "error_message",
    "collected_at",
)


def _collection(
    product_id: str,
    *,
    status: ProductStatus = ProductStatus.SUCCESS,
    error_message: str = "",
) -> ProductCollection:
    return ProductCollection.from_records(
        [
            ProductRecord.success_fixture(
                product_id, status=status, error_message=error_message
            )
        ],
        generated_at="2026-07-16T20:00:00+08:00",
        blocked=False,
    )


def _write_baseline(target: Path, product_id: str) -> None:
    ProductOutputBundle(target).write(_collection(product_id))


def test_verify_bundle_accepts_matching_outputs(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    assert verify_product_bundle(target) == {
        "products": 1,
        "unique_ids": 1,
        "success": 1,
        "partial": 0,
        "failed": 0,
    }


def test_verify_bundle_rejects_json_id_mismatch(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    payload = json.loads((target / "products.json").read_text(encoding="utf-8"))
    payload["products"][0]["product_id"] = "other"
    (target / "products.json").write_text(
        json.dumps(payload), encoding="utf-8"
    )
    with pytest.raises(ProductBundleContractError, match="product ids"):
        verify_product_bundle(target)


@pytest.mark.parametrize("missing", list(BUNDLE_FILES))
def test_verify_bundle_rejects_missing_artifact(tmp_path: Path, missing: str) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    (target / missing).unlink()
    with pytest.raises(ProductBundleContractError, match="missing"):
        verify_product_bundle(target)


def test_verify_bundle_rejects_wrong_excel_columns(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["wrong", "columns"])
    workbook.save(target / "products.xlsx")
    with pytest.raises(ProductBundleContractError, match="columns"):
        verify_product_bundle(target)


def test_verify_bundle_rejects_wrong_schema_version(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    payload = json.loads((target / "products.json").read_text(encoding="utf-8"))
    payload["schema_version"] = 2
    (target / "products.json").write_text(
        json.dumps(payload), encoding="utf-8"
    )
    with pytest.raises(ProductBundleContractError, match="schema_version"):
        verify_product_bundle(target)


def test_verify_bundle_rejects_duplicate_product_ids(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(EXCEL_COLUMNS))
    sheet.append([
        "1", "web-scraping.dev", "https://web-scraping.dev/product/1",
        "Product 1", "", "", "", 9.99, None, "USD", "", 0,
        "SUCCESS", "", "2026-07-16T20:00:00+08:00",
    ])
    sheet.append([
        "1", "web-scraping.dev", "https://web-scraping.dev/product/1",
        "Product 1", "", "", "", 9.99, None, "USD", "", 0,
        "SUCCESS", "", "2026-07-16T20:00:00+08:00",
    ])
    workbook.save(target / "products.xlsx")
    with pytest.raises(ProductBundleContractError, match="unique"):
        verify_product_bundle(target)


def test_verify_bundle_rejects_more_than_ten_products(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(EXCEL_COLUMNS))
    for index in range(11):
        sheet.append([
            f"prod-{index}", "web-scraping.dev",
            f"https://web-scraping.dev/product/{index}", "Product", "", "",
            "", 9.99, None, "USD", "", 0, "SUCCESS", "",
            "2026-07-16T20:00:00+08:00",
        ])
    workbook.save(target / "products.xlsx")
    with pytest.raises(ProductBundleContractError, match="1 and 10"):
        verify_product_bundle(target)


def test_verify_bundle_rejects_missing_collected_at(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(EXCEL_COLUMNS))
    sheet.append([
        "1", "web-scraping.dev", "https://web-scraping.dev/product/1",
        "Product 1", "", "", "", 9.99, None, "USD", "", 0,
        "SUCCESS", "", "",
    ])
    workbook.save(target / "products.xlsx")
    with pytest.raises(ProductBundleContractError, match="collected_at"):
        verify_product_bundle(target)


def test_verify_bundle_rejects_html_without_embedded_id(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    (target / "gallery.html").write_text(
        "<html><body>no embedded payload</body></html>",
        encoding="utf-8",
    )
    with pytest.raises(ProductBundleContractError, match="embedded"):
        verify_product_bundle(target)


@pytest.mark.parametrize(
    "forbidden_token",
    ["<script src=", "fetch(", "@import url("],
)
def test_verify_bundle_rejects_html_with_external_dependency(
    tmp_path: Path, forbidden_token: str
) -> None:
    target = tmp_path / "demo"
    _write_baseline(target, "1")
    html = (target / "gallery.html").read_text(encoding="utf-8")
    (target / "gallery.html").write_text(
        html + f"\n{forbidden_token}\n", encoding="utf-8"
    )
    with pytest.raises(ProductBundleContractError, match="external"):
        verify_product_bundle(target)


def test_verify_bundle_counts_partial_and_failed_records(tmp_path: Path) -> None:
    target = tmp_path / "demo"
    ProductOutputBundle(target).write(
        ProductCollection.from_records(
            [
                ProductRecord.success_fixture("1"),
                ProductRecord.success_fixture(
                    "2",
                    status=ProductStatus.PARTIAL,
                    error_message="brand missing",
                ),
                ProductRecord.failure(
                    ProductListing(
                        "3",
                        "https://web-scraping.dev/product/3",
                        "",
                    ),
                    ProductStatus.NETWORK_ERROR,
                    "offline",
                ),
            ],
            generated_at="2026-07-16T20:00:00+08:00",
            blocked=False,
        )
    )
    assert verify_product_bundle(target) == {
        "products": 3,
        "unique_ids": 3,
        "success": 1,
        "partial": 1,
        "failed": 1,
    }


def test_verify_bundle_accepts_runner_generated_outputs(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """End-to-end: runner -> output bundle -> verifier contract check.

    Drives :class:`ProductRunner` with a fake adapter that surfaces
    more listings than the run limit and a non-None ``next_url`` so
    the discovery cap is enforced before the second page is fetched.
    A logger is attached to keep the runner's success-path logging
    within scope of the verifier integration. The runner-paced
    ``time.sleep`` calls are patched out so the test stays offline
    and fast.
    """
    monkeypatch.setattr("app.product_runner.time.sleep", lambda _seconds: None)
    monkeypatch.setattr("app.product_runner.time.monotonic", lambda: 0.0)

    listings = (
        ProductListing("1", "https://web-scraping.dev/product/1"),
        ProductListing("2", "https://web-scraping.dev/product/2"),
    )
    records_by_id = {
        pid: ProductRecord.success_fixture(pid) for pid in ("1", "2")
    }

    class _FakeAdapter:
        PRODUCTS_URL = "https://web-scraping.dev/products"

        def __init__(self) -> None:
            self.list_calls: list[str] = []
            self.product_calls: list[str] = []

        def fetch_products_page(self, _tab, url: str) -> ProductPage:
            self.list_calls.append(url)
            return ProductPage(listings, "https://web-scraping.dev/products?page=2")

        def fetch_product(self, _tab, listing: ProductListing) -> ProductRecord:
            self.product_calls.append(listing.product_id)
            return records_by_id[listing.product_id]

    adapter = _FakeAdapter()
    collection = ProductRunner(
        adapter,
        object(),
        max_products=1,
        min_interval_seconds=0,
        logger=logging.getLogger("test_verify_products.runner"),
    ).run()

    assert [record.product_id for record in collection.records] == ["1"]
    assert adapter.product_calls == ["1"]
    # Discovery saw two pages' worth of listings, but the runner's
    # discovery cap stopped at the first listing because the run
    # limit is 1 and the second page was never fetched.
    assert adapter.list_calls == ["https://web-scraping.dev/products"]

    target = tmp_path / "demo"
    ProductOutputBundle(target).write(collection)
    assert verify_product_bundle(target) == {
        "products": 1,
        "unique_ids": 1,
        "success": 1,
        "partial": 0,
        "failed": 0,
    }
